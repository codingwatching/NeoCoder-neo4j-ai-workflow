#!/usr/bin/env python3
"""
Excel File Processor for NeoCoder Data Analysis

This script provides specialized handling for Excel files (.xlsx, .xls),
including multi-sheet extraction, metadata preservation, and format conversion.

Author: NeoCoder Data Analysis Team
Created: 2025
"""

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

# Add the src directory to Python path if running directly
if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[2]
    sys.path.insert(0, str(project_root / "src"))

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("excel_processing.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


class ExcelProcessor:
    """
    Specialized Excel file processor.

    Features:
    - Multi-sheet extraction and conversion
    - Metadata preservation and analysis
    - Cell formatting detection
    - Data type analysis per sheet
    - Smart column naming
    - Chart and formula detection
    """

    def __init__(self, input_file: str, output_dir: Optional[str] = None):
        """
        Initialize the Excel processor.

        Args:
            input_file: Path to Excel file
            output_dir: Output directory for converted files
        """
        self.input_file = Path(input_file)
        self.output_dir = Path(output_dir) if output_dir else self.input_file.parent

        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Processing results
        self.results: Dict[str, Any] = {
            "file_info": {},
            "sheets": [],
            "metadata": {},
            "conversion_summary": {},
        }

        # Check if pandas and openpyxl are available
        import importlib.util

        if importlib.util.find_spec("pandas") and importlib.util.find_spec("openpyxl"):
            self.pandas_available = True
            logger.info("Excel processing libraries available")
        else:
            self.pandas_available = False
            logger.error("Required libraries (pandas, openpyxl) not available")
            logger.error("Install with: pip install pandas openpyxl xlrd")

        logger.info("ExcelProcessor initialized")
        logger.info(f"Input file: {self.input_file}")
        logger.info(f"Output directory: {self.output_dir}")

    def analyze_file_info(self) -> Dict[str, Any]:
        """
        Analyze basic Excel file information.

        Returns:
            File information dictionary
        """
        if not self.input_file.exists():
            raise FileNotFoundError(f"Excel file not found: {self.input_file}")

        file_stats = self.input_file.stat()

        file_info = {
            "filename": self.input_file.name,
            "file_path": str(self.input_file),
            "file_size": file_stats.st_size,
            "file_size_mb": round(file_stats.st_size / (1024 * 1024), 2),
            "created_time": datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
            "modified_time": datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
            "extension": self.input_file.suffix.lower(),
        }

        logger.info(
            f"File analysis: {file_info['file_size_mb']} MB, {file_info['extension']} format"
        )
        return file_info

    def get_workbook_metadata(self) -> Dict[str, Any]:
        """
        Extract workbook-level metadata using openpyxl.

        Returns:
            Workbook metadata dictionary
        """
        if not self.pandas_available:
            return {"error": "Required libraries not available"}

        try:
            import openpyxl

            # Load workbook for metadata
            wb = openpyxl.load_workbook(self.input_file, data_only=False)

            metadata = {
                "sheet_names": wb.sheetnames,
                "sheet_count": len(wb.sheetnames),
                "properties": {},
                "defined_names": [],
                "charts_detected": False,
                "formulas_detected": False,
            }

            # Get document properties
            if wb.properties:
                props = wb.properties
                metadata["properties"] = {
                    "title": props.title,
                    "creator": props.creator,
                    "subject": props.subject,
                    "description": props.description,
                    "keywords": props.keywords,
                    "category": props.category,
                    "created": props.created.isoformat() if props.created else None,
                    "modified": props.modified.isoformat() if props.modified else None,
                    "version": props.version,
                }

            # Check for defined names (named ranges)
            if wb.defined_names:
                metadata["defined_names"] = [
                    {"name": name.name, "value": str(name.value)}
                    for name in wb.defined_names.definedName
                ]

            # Quick scan for charts and formulas
            for sheet in wb.worksheets:
                # Check for charts using a safer approach
                try:
                    # Try to access the charts attribute if it exists
                    charts_attr = getattr(sheet, "charts", None)
                    if charts_attr is not None and len(charts_attr) > 0:
                        metadata["charts_detected"] = True
                except (AttributeError, TypeError):
                    # Skip chart detection if attributes are not accessible
                    pass

                # Sample some cells for formulas
                formula_count = 0
                cell_count = 0

                for row in sheet.iter_rows(
                    max_row=min(100, sheet.max_row), max_col=min(20, sheet.max_column)
                ):
                    for cell in row:
                        if cell.value is not None:
                            cell_count += 1
                            if cell.data_type == "f":  # Formula
                                formula_count += 1

                        if cell_count >= 200:  # Sample limit
                            break
                    if cell_count >= 200:
                        break

                if formula_count > 0:
                    metadata["formulas_detected"] = True
                    break

            wb.close()
            logger.info(
                f"Metadata extracted: {metadata['sheet_count']} sheets, charts: {metadata['charts_detected']}, formulas: {metadata['formulas_detected']}"
            )

            return metadata

        except Exception as e:
            logger.error(f"Error extracting metadata: {e}")
            return {"error": str(e)}

    def analyze_sheet_data(self, sheet_name: str, df: Any) -> Dict[str, Any]:
        """
        Analyze data in a specific sheet.

        Args:
            sheet_name: Name of the sheet
            df: Pandas DataFrame with sheet data

        Returns:
            Sheet analysis dictionary
        """
        import pandas as pd

        analysis: Dict[str, Any] = {
            "sheet_name": sheet_name,
            "dimensions": {"rows": len(df), "columns": len(df.columns)},
            "columns": {},
            "data_quality": {},
            "summary_stats": {},
        }

        # Analyze each column
        for col in df.columns:
            col_data = df[col]

            # Basic info
            col_info: Dict[str, Any] = {
                "name": col,
                "non_null_count": int(col_data.count()),
                "null_count": int(col_data.isnull().sum()),
                "null_percentage": float(
                    (col_data.isnull().sum() / len(col_data)) * 100
                ),
                "unique_count": int(col_data.nunique()),
                "data_type": str(col_data.dtype),
            }

            # Sample values (first 5 non-null)
            sample_values = col_data.dropna().head(5).tolist()
            col_info["sample_values"] = [str(v) for v in sample_values]

            # Detect actual data type
            non_null_data = col_data.dropna()

            if len(non_null_data) > 0:
                # Check for numeric data
                try:
                    numeric_data = pd.to_numeric(non_null_data, errors="coerce")
                    # Count valid numeric values after coercion
                    numeric_count = numeric_data.count()

                    if numeric_count / len(non_null_data) > 0.8:
                        col_info["detected_type"] = "numeric"

                        # Calculate stats for numeric columns
                        col_info["stats"] = {
                            "mean": float(numeric_data.mean()),
                            "median": float(numeric_data.median()),
                            "min": float(numeric_data.min()),
                            "max": float(numeric_data.max()),
                            "std": (
                                float(numeric_data.std())
                                if len(numeric_data) > 1
                                else 0.0
                            ),
                        }

                        # Check if integers
                        if all(x.is_integer() for x in numeric_data if pd.notna(x)):
                            col_info["detected_type"] = "integer"

                except (ValueError, TypeError) as e:
                    logger.debug(f"Column {col} is not numeric: {e}")

                # Check for dates
                if col_info.get("detected_type") != "numeric":
                    try:
                        date_data = pd.to_datetime(non_null_data, errors="coerce")
                        date_count = date_data.count()

                        if date_count / len(non_null_data) > 0.7:
                            col_info["detected_type"] = "datetime"

                            col_info["stats"] = {
                                "earliest": str(date_data.min()),
                                "latest": str(date_data.max()),
                                "date_range_days": (
                                    date_data.max() - date_data.min()
                                ).days,
                            }
                    except (ValueError, TypeError) as e:
                        logger.debug(f"Column {col} is not datetime: {e}")

                # Default to categorical/text
                if "detected_type" not in col_info:
                    if col_info["unique_count"] <= min(20, len(non_null_data) * 0.5):
                        col_info["detected_type"] = "categorical"

                        # Get value counts for categorical
                        value_counts = col_data.value_counts().head(10)
                        col_info["value_counts"] = {
                            str(k): int(v) for k, v in value_counts.items()
                        }
                    else:
                        col_info["detected_type"] = "text"

            analysis["columns"][col] = col_info

        # Overall data quality assessment
        total_cells = len(df) * len(df.columns)
        missing_cells = df.isnull().sum().sum()

        analysis["data_quality"] = {
            "total_cells": int(total_cells),
            "missing_cells": int(missing_cells),
            "completeness_percentage": (
                ((total_cells - missing_cells) / total_cells) * 100
                if total_cells > 0
                else 0
            ),
            "has_duplicates": df.duplicated().any(),
            "duplicate_count": int(df.duplicated().sum()),
        }

        # Summary statistics
        numeric_columns = [
            col
            for col, info in analysis["columns"].items()
            if info.get("detected_type") in ["numeric", "integer"]
        ]

        if numeric_columns:
            analysis["summary_stats"]["numeric_columns"] = len(numeric_columns)
            analysis["summary_stats"]["total_numeric_values"] = int(
                df[numeric_columns].count().sum()
            )

        categorical_columns = [
            col
            for col, info in analysis["columns"].items()
            if info.get("detected_type") == "categorical"
        ]

        if categorical_columns:
            analysis["summary_stats"]["categorical_columns"] = len(categorical_columns)

        logger.info(
            f"Sheet '{sheet_name}' analyzed: {analysis['dimensions']['rows']} rows, "
            f"{analysis['dimensions']['columns']} columns, "
            f"{analysis['data_quality']['completeness_percentage']:.1f}% complete"
        )

        return analysis

    def convert_sheet_to_csv(
        self, sheet_name: str, df: Any, base_filename: str
    ) -> Dict[str, Any]:
        """
        Convert a sheet to CSV format.

        Args:
            sheet_name: Name of the sheet
            df: Pandas DataFrame with sheet data
            base_filename: Base filename for output

        Returns:
            Conversion result dictionary
        """
        # Clean sheet name for filename
        clean_sheet_name = "".join(
            c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")
        ).strip()
        clean_sheet_name = clean_sheet_name.replace(" ", "_")

        # Generate output filename
        if clean_sheet_name.lower() == "sheet1" or not clean_sheet_name:
            output_file = self.output_dir / f"{base_filename}.csv"
        else:
            output_file = self.output_dir / f"{base_filename}_{clean_sheet_name}.csv"

        try:
            # Save to CSV
            df.to_csv(output_file, index=False, encoding="utf-8")

            result = {
                "sheet_name": sheet_name,
                "output_file": str(output_file),
                "output_filename": output_file.name,
                "rows_exported": len(df),
                "columns_exported": len(df.columns),
                "status": "success",
            }

            logger.info(f"Sheet '{sheet_name}' converted to {output_file.name}")
            return result

        except Exception as e:
            logger.error(f"Error converting sheet '{sheet_name}': {e}")
            return {"sheet_name": sheet_name, "status": "error", "error": str(e)}

    def process_excel_file(self) -> Dict[str, Any]:
        """
        Process the Excel file completely.

        Returns:
            Complete processing results
        """
        if not self.pandas_available:
            return {
                "status": "error",
                "error": "Required libraries (pandas, openpyxl) not available",
            }

        logger.info(f"Starting Excel file processing: {self.input_file.name}")

        try:
            import pandas as pd

            # Analyze file info
            self.results["file_info"] = self.analyze_file_info()

            # Get metadata
            self.results["metadata"] = self.get_workbook_metadata()

            # Read all sheets
            try:
                excel_data = pd.read_excel(
                    self.input_file,
                    sheet_name=None,  # Read all sheets
                    engine=(
                        "openpyxl"
                        if self.input_file.suffix.lower() == ".xlsx"
                        else "xlrd"
                    ),
                )
            except Exception as e:
                logger.error(f"Error reading Excel file: {e}")
                return {"status": "error", "error": f"Could not read Excel file: {e}"}

            if not excel_data:
                return {"status": "warning", "message": "No sheets found in Excel file"}

            # Process each sheet
            base_filename = self.input_file.stem
            sheet_results = []
            successful_conversions = 0

            for sheet_name, df in excel_data.items():
                logger.info(f"Processing sheet: {sheet_name}")

                # Skip empty sheets
                if df.empty:
                    logger.warning(f"Sheet '{sheet_name}' is empty, skipping")
                    sheet_results.append(
                        {
                            "sheet_name": sheet_name,
                            "status": "skipped",
                            "reason": "empty",
                        }
                    )
                    continue

                # Analyze sheet data
                sheet_analysis = self.analyze_sheet_data(sheet_name, df)

                # Convert to CSV
                conversion_result = self.convert_sheet_to_csv(
                    sheet_name, df, base_filename
                )

                # Combine results
                sheet_result = {**sheet_analysis, "conversion": conversion_result}

                sheet_results.append(sheet_result)

                if conversion_result["status"] == "success":
                    successful_conversions += 1

            self.results["sheets"] = sheet_results

            # Create conversion summary
            self.results["conversion_summary"] = {
                "total_sheets": len(excel_data),
                "processed_sheets": len(
                    [
                        s
                        for s in sheet_results
                        if s.get("conversion", {}).get("status") == "success"
                    ]
                ),
                "skipped_sheets": len(
                    [s for s in sheet_results if s.get("status") == "skipped"]
                ),
                "failed_sheets": len(
                    [
                        s
                        for s in sheet_results
                        if s.get("conversion", {}).get("status") == "error"
                    ]
                ),
                "success_rate": (
                    (successful_conversions / len(excel_data)) * 100
                    if excel_data
                    else 0
                ),
            }

            logger.info(
                f"Excel processing completed: {successful_conversions}/{len(excel_data)} sheets converted successfully"
            )

            return {"status": "success", "results": self.results}

        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return {"status": "error", "error": str(e)}

    def generate_report(self, processing_results: Dict) -> str:
        """
        Generate a comprehensive processing report.

        Args:
            processing_results: Results from process_excel_file()

        Returns:
            Formatted report string
        """
        report = []
        report.append("# Excel File Processing Report")
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")

        if processing_results["status"] != "success":
            report.append("## ❌ Processing Failed")
            report.append(
                f"**Error:** {processing_results.get('error', 'Unknown error')}"
            )
            return "\n".join(report)

        results = processing_results["results"]

        # File information
        file_info = results["file_info"]
        report.append("## 📄 File Information")
        report.append(f"- **Filename:** {file_info['filename']}")
        report.append(f"- **File Size:** {file_info['file_size_mb']} MB")
        report.append(f"- **Format:** {file_info['extension'].upper()}")
        report.append(f"- **Modified:** {file_info['modified_time'][:19]}")
        report.append("")

        # Metadata
        metadata = results["metadata"]
        if metadata and "error" not in metadata:
            report.append("## 📊 Workbook Metadata")
            report.append(f"- **Total Sheets:** {metadata['sheet_count']}")
            report.append(f"- **Sheet Names:** {', '.join(metadata['sheet_names'])}")

            if metadata.get("charts_detected"):
                report.append("- **Charts Detected:** ✅")

            if metadata.get("formulas_detected"):
                report.append("- **Formulas Detected:** ✅")

            # Document properties
            props = metadata.get("properties", {})
            if any(props.values()):
                report.append("- **Document Properties:**")
                for prop, value in props.items():
                    if value:
                        report.append(f"  - {prop.title()}: {value}")

            report.append("")

        # Conversion summary
        summary = results["conversion_summary"]
        report.append("## 🔄 Conversion Summary")
        report.append(f"- **Total Sheets:** {summary['total_sheets']}")
        report.append(f"- **Successfully Converted:** {summary['processed_sheets']}")
        report.append(f"- **Skipped (Empty):** {summary['skipped_sheets']}")
        report.append(f"- **Failed:** {summary['failed_sheets']}")
        report.append(f"- **Success Rate:** {summary['success_rate']:.1f}%")
        report.append("")

        # Individual sheet details
        report.append("## 📋 Sheet Details")
        report.append("")

        sheets = results["sheets"]
        for sheet in sheets:
            if sheet.get("status") == "skipped":
                report.append(f"### ⏭️ {sheet['sheet_name']} (Skipped)")
                report.append(f"**Reason:** {sheet.get('reason', 'Unknown')}")
                report.append("")
                continue

            conversion = sheet.get("conversion", {})
            status_icon = "✅" if conversion.get("status") == "success" else "❌"

            report.append(f"### {status_icon} {sheet['sheet_name']}")

            if conversion.get("status") == "success":
                report.append(f"- **Output File:** {conversion['output_filename']}")
                report.append(
                    f"- **Dimensions:** {sheet['dimensions']['rows']:,} rows × {sheet['dimensions']['columns']} columns"
                )

                # Data quality
                quality = sheet["data_quality"]
                report.append(
                    f"- **Completeness:** {quality['completeness_percentage']:.1f}%"
                )

                if quality["has_duplicates"]:
                    report.append(
                        f"- **Duplicates:** {quality['duplicate_count']} rows"
                    )

                # Column types summary
                column_types: Dict[str, int] = {}
                for col_info in sheet["columns"].values():
                    detected_type = col_info.get("detected_type", "unknown")
                    column_types[detected_type] = column_types.get(detected_type, 0) + 1

                if column_types:
                    type_summary = ", ".join(
                        [
                            f"{count} {type_name}"
                            for type_name, count in column_types.items()
                        ]
                    )
                    report.append(f"- **Column Types:** {type_summary}")

            else:
                report.append(
                    f"- **Error:** {conversion.get('error', 'Unknown error')}"
                )

            report.append("")

        # Ready for analysis
        successful_files = [
            sheet["conversion"]
            for sheet in sheets
            if sheet.get("conversion", {}).get("status") == "success"
        ]

        if successful_files:
            report.append("## 🚀 Ready for Data Analysis")
            report.append("")
            report.append("The following CSV files are ready for analysis:")
            report.append("")

            for conversion in successful_files:
                output_file = conversion["output_file"]
                filename = conversion["output_filename"]

                report.append(f"### {filename}")
                report.append("")
                report.append("```python")
                report.append("# Load dataset for analysis")
                report.append("load_dataset(")
                report.append(f'    file_path="{output_file}",')
                report.append(f'    dataset_name="{Path(filename).stem}",')
                report.append('    source_type="csv"')
                report.append(")")
                report.append("")
                report.append("# Generate insights")
                report.append('generate_insights(dataset_id="DATASET_ID")')
                report.append("```")
                report.append("")

        # Recommendations
        report.append("## 💡 Recommendations")
        report.append("")

        if summary["success_rate"] == 100:
            report.append("✅ All sheets converted successfully!")
        elif summary["success_rate"] >= 75:
            report.append("✅ Most sheets converted successfully")
        elif summary["failed_sheets"] > 0:
            report.append(
                "⚠️ Some sheets failed to convert - check error messages above"
            )

        if metadata.get("formulas_detected"):
            report.append(
                "📝 Formulas were detected - converted values may differ from original calculations"
            )

        if metadata.get("charts_detected"):
            report.append(
                "📈 Charts were detected but not converted - original visualization lost"
            )

        # Check for data quality issues
        total_completeness = sum(
            sheet["data_quality"]["completeness_percentage"]
            for sheet in sheets
            if "data_quality" in sheet
        )
        avg_completeness = (
            total_completeness / len([s for s in sheets if "data_quality" in s])
            if sheets
            else 0
        )

        if avg_completeness < 90:
            report.append("⚠️ Data completeness is below 90% - consider data cleaning")

        report.append("")

        return "\n".join(report)

    def save_results(self, processing_results: Dict) -> Path:
        """
        Save processing results to JSON file.

        Args:
            processing_results: Results from process_excel_file()

        Returns:
            Path to saved results file
        """
        results_file = (
            self.output_dir / f"{self.input_file.stem}_excel_processing_results.json"
        )

        with open(results_file, "w", encoding="utf-8") as f:
            json.dump(processing_results, f, indent=2, default=str)

        logger.info(f"Processing results saved: {results_file}")
        return results_file


def main() -> int:
    """Main function for command line usage."""
    parser = argparse.ArgumentParser(
        description="Process Excel files for NeoCoder data analysis"
    )
    parser.add_argument("input_file", help="Input Excel file path (.xlsx or .xls)")
    parser.add_argument("-o", "--output-dir", help="Output directory for CSV files")
    parser.add_argument(
        "--save-results",
        action="store_true",
        help="Save detailed processing results to JSON",
    )

    args = parser.parse_args()

    # Validate input file
    input_path = Path(args.input_file)
    if not input_path.exists():
        logger.error(f"Input file not found: {input_path}")
        return 1

    if input_path.suffix.lower() not in [".xlsx", ".xls"]:
        logger.error(f"Input file must be Excel format (.xlsx or .xls): {input_path}")
        return 1

    # Initialize processor
    processor = ExcelProcessor(input_file=str(input_path), output_dir=args.output_dir)

    # Process file
    results = processor.process_excel_file()

    # Generate and display report
    report = processor.generate_report(results)
    print(report)

    # Save detailed results if requested
    if args.save_results and results["status"] == "success":
        processor.save_results(results)

    # Save report
    if results["status"] == "success":
        report_file = processor.output_dir / f"{input_path.stem}_excel_report.md"
        with open(report_file, "w", encoding="utf-8") as f:
            f.write(report)
        logger.info(f"Excel processing report saved: {report_file}")

    # Return appropriate exit code
    if results["status"] == "success":
        summary = results["results"]["conversion_summary"]
        if summary["failed_sheets"] == 0:
            return 0  # All successful
        elif summary["processed_sheets"] > 0:
            return 2  # Partial success
        else:
            return 1  # All failed
    else:
        return 1


if __name__ == "__main__":
    exit(main())
